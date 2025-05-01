import base64
import streamlit as st
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from copy import copy
from datetime import datetime
from io import BytesIO
import pandas as pd
from fpdf import FPDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
# ---------- Helper Functions ----------
def clear_cell_style(cell):
    cell.font = copy(NamedStyle().font)
    cell.fill = copy(NamedStyle().fill)
    cell.border = copy(NamedStyle().border)
    cell.alignment = copy(NamedStyle().alignment)
    cell.number_format = NamedStyle().number_format
    cell.protection = copy(NamedStyle().protection)

def load_and_process(file_path, sheet_index,month_string):
    wb = load_workbook(file_path)
    ws = wb.worksheets[sheet_index]
    max_row = ws.max_row
    max_col = ws.max_column

    print_columns = []
    for row in range(1, max_row + 1):
        for col in range(2, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip().lower() == "print":
                print_columns.append((row, col))

    if not print_columns:
        return wb, ws  # No processing needed

    # First print and task columns become destination
    dest_header_row, dest_print_col = print_columns[0]
    dest_task_col = dest_print_col - 1
    all_entries = []
    above_val=[]
    flag=0
    header_dict={}
    # Collect from all print columns
    for header_row, print_col in print_columns:
        task_col = print_col - 1
        above_entries=[]
        header_entries=[]
        # Check the content in the cells above the task column header
        for row in range(1, header_row):  # Iterate from row 1 to the row above the header
            task_cell = ws.cell(row=row, column=task_col)
            # if task_cell.value:  # If there's any value in the cell
            if "month:" in str(task_cell.value).lower():
                task_cell.value = "MONTH: " + month_string

            above_entries.append((task_cell.value, task_cell))
        above_val.append(above_entries)
        header_cell=ws.cell(row=header_row,column=task_col)
        header_entries.append((header_cell.value, copy(header_cell._style) if header_cell.has_style else None))
        if(flag==1):
            header_dict[task_col]=header_entries
        flag=1
                

    # Collect from all print columns
    for header_row, print_col in print_columns:
        task_col = print_col - 1
        if task_col in header_dict:
            # print(header_dict)
            all_entries=all_entries+header_dict[task_col]
        for row in range(header_row + 1, max_row + 1):
            task_cell = ws.cell(row=row, column=task_col)
            print_cell = ws.cell(row=row, column=print_col)
            if str(print_cell.value).lower() == "p":
                all_entries.append((task_cell.value, copy(task_cell._style) if task_cell.has_style else None))
    print(all_entries)
    # Clear destination column
    for row in range(dest_header_row + 1, max_row + 1):
        clear_cell_style(ws.cell(row=row, column=dest_task_col))
        ws.cell(row=row, column=dest_task_col).value = None

    
    # Write consolidated tasks into destination task column
    for idx, task_data in enumerate(all_entries):  # Removed the tuple unpacking here
        task_value, task_style = task_data  # Explicitly unpack the tuple
        new_row = dest_header_row + 1 + idx
        task_cell = ws.cell(row=new_row, column=dest_task_col)
        task_cell.value = task_value or " "
        if task_style:
            task_cell._style = copy(task_style)

    # Delete all columns except the destination task column
    for col in reversed(range(1, max_col + 1)):
        if col != dest_task_col:
            ws.delete_cols(col)


    return wb, ws,above_val




def save_workbook(wb, filename):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



def sheet_to_pdf(ws, pdf_path,above_val):
    # print(above_val)
    c = canvas.Canvas(pdf_path, pagesize=A4)
    width, height = A4
    margin = 15 * mm
    col_width = (width - 2 * margin) / 2
    row_height = 8 * mm
    text_offset_y = 2 * mm
    max_rows_per_col = int((height - 2 * margin) / row_height)-1
    heading_reserved_height = 3 * row_height  # space reserved for heading


    # Extract entries
    entries = []
    headings=[]
    
    for row in ws.iter_rows(min_row=2, max_col=1):  # skip header
        cell = row[0]
        if cell.value is not None:
            font = cell.font
            fill = cell.fill  # <-- get fill object
            text = str(cell.value)
            size = font.sz or 10
            bold = font.bold

            # Text color
            color = "#000000"
            if (cell.value!=" ") and font.color and hasattr(font.color, "rgb") and font.color.rgb:
                color = f"#{font.color.rgb[2:]}"  # skip 'FF' alpha

            # Background color
            bg_color = "#FFFFFF"  # default white
            if fill and fill.start_color:
                fill_color = fill.start_color

                if fill_color.type == "rgb" and fill_color.rgb:
                    # Direct RGB value, remove 'FF' alpha if present
                    bg_color = f"#FFFFFF"  # last 6 characters: RRGGBB

                elif fill_color.type == "theme":
                    # Theme-based color — hard to resolve without mapping, fallback
                    bg_color = "#DDD9C4"  # or set a placeholder like "#D9E1F2" for Excel default theme

                elif fill_color.type == "indexed":
                    # Indexed color (rare), can try mapping from indexed table
                    index = fill_color.indexed
                    # You can build a mapping for known index values if needed
                    bg_color = f"indexed_{index}"

            entries.append({
                "text": text,
                "size": size,
                "bold": bold,
                "color": color,
                "background": bg_color  # add background color info
            })

    
    for row in above_val:  # skip header
        # print(row)
        cell = row[0][1]
        if cell.value is not None:
            font = cell.font
            text = str(cell.value)
            size = font.sz or 10
            bold = font.bold
            color = "#000000"
            if font.color and hasattr(font.color, "rgb") and font.color.rgb:
                color = f"#{font.color.rgb[2:]}"
            headings.append({
                "text": text,
                "size": size,
                "bold": bold,
                "color": color
            })

    idx = 0
    total = len(entries)

    while idx < total:
        y = height - margin
        # --- Heading ---
       
        start_x = (width) / 2
        x = start_x-150
        for entry in headings:
            
            c.setFont("Helvetica-Bold" if entry["bold"] else "Helvetica", entry["size"])
            c.setFillColor(HexColor(entry["color"]))
            c.drawString(x, y - text_offset_y, entry["text"])
            x +=  200
        
        

        y -= heading_reserved_height  # reserve space after heading
        c.setFillColor(HexColor("#D9D9D9"))
        c.rect(margin-2, y+45, col_width*2, row_height-20, stroke=0, fill=1)

        # --- Left column ---
        left_y = y
        for i in range(max_rows_per_col):
            if idx >= total:
                break
            entry = entries[idx]
            # print(entry)
            x = margin
            c.setFillColor(HexColor(entry["background"]))
            c.rect(x - 2, left_y - text_offset_y - 2, col_width, row_height-2, stroke=0, fill=1)
            c.setFont("Helvetica-Bold" if entry["bold"] else "Helvetica", entry["size"])
            c.setFillColor(HexColor(entry["color"]))
            c.drawString(x, left_y - text_offset_y, entry["text"])
            left_y -= row_height
            idx += 1

        # --- Right column ---
        right_y = y
        for i in range(max_rows_per_col):
            if idx >= total:
                break
            entry = entries[idx]
            x = margin + col_width
            c.setFillColor(HexColor(entry["background"]))
            c.rect(x - 2, right_y - text_offset_y - 2, col_width, row_height-2, stroke=0, fill=1)
            c.setFont("Helvetica-Bold" if entry["bold"] else "Helvetica", entry["size"])
            c.setFillColor(HexColor(entry["color"]))
            c.drawString(x, right_y - text_offset_y, entry["text"])
            right_y -= row_height
            idx += 1

        c.showPage()

    c.save()

# ---------- Streamlit UI ----------
st.title("📊 MONTHLY MANAGER")

today_str = datetime.today().strftime('%Y-%m-%d')
task_folder = "Task Files"
modified_folder = os.path.join("Modified Files", today_str)
os.makedirs(modified_folder, exist_ok=True)

files = [f for f in os.listdir(task_folder) if f.endswith(".xlsx")]

if not files:
    st.warning("No Excel files found in 'task files' folder.")
else:
    selected_file = st.selectbox("Select Excel file", files)
    file_path = os.path.join(task_folder, selected_file)

    # Load workbook and list sheets
    wb_preview = load_workbook(file_path, read_only=True)
    sheet_names = wb_preview.sheetnames
    selected_sheet = st.selectbox("Select Sheet", sheet_names)
    sheet_index = sheet_names.index(selected_sheet)
    
    now = datetime.now()
    current_year = now.year
    current_month = now.strftime("%B")

    # Allow user to type year
    selected_year = st.number_input("Enter Year", value=current_year, min_value=1900, max_value=2100, step=1)

    # Allow user to select month
    months = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    selected_month = st.selectbox("Select Month", months, index=months.index(current_month))

    # Display selection
    st.write(f"Selected Month: **{selected_month} {int(selected_year)}**")
    month_string = f"{selected_month.upper()}-{str(selected_year)[-2:]}"
    st.write("Formatted String:", month_string)     

    if st.button("🔄 Process Selected Sheet"):
        wb_processed, ws_processed,above_val = load_and_process(file_path, sheet_index,month_string)
        wb_preview.close()

        # Save Excel
        filename_base = os.path.splitext(selected_file)[0]
        output_excel_name = f"{filename_base}_{selected_sheet}_{today_str}.xlsx"
        output_excel_path = os.path.join(modified_folder, output_excel_name)
        wb_processed.save(output_excel_path)

        # Create PDF
        output_pdf_name = f"{filename_base}_{selected_sheet}_{today_str}.pdf"
        output_pdf_path = os.path.join(modified_folder, output_pdf_name)
        sheet_to_pdf(ws_processed, output_pdf_path,above_val)

        st.success("✅ Sheet processed, Excel & PDF saved!")
            
        if output_pdf_path:
                    with open(output_pdf_path, "rb") as f:
                        base64_pdf = base64.b64encode(f.read()).decode('utf-8')
                        pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="700" height="500" type="application/pdf"></iframe>'
                    st.markdown(pdf_display, unsafe_allow_html=True)
